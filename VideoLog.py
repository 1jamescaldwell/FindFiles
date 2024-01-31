def highlight_excel(run_row,run_column,color):

    import os
    import pandas as pd

    # Path to the root folder
    root_folder = r'\\cab-fs07.mae.virginia.edu\NewData\FieldTurf\2023-FieldTurf\1Video'

    # Path to the Excel file
    excel_file_path = r'\\cab-fs07.mae.virginia.edu\NewData\FieldTurf\2023-FieldTurf\1Video\Video Log Highlighted.xlsx'

    # Load the Excel file into a pandas DataFrame
    video_log = pd.read_excel(excel_file_path)

    file_count = 0

    import openpyxl
    from openpyxl.styles import PatternFill

    # Load the workbook
    # excel_file_path = r'path\to\your\excel_file.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the first sheet
    sheet = workbook.active

    # Specify the cell to highlight (2nd row, 1st column)
    # cell_to_highlight = sheet.cell(row=2, column=1)
    cell_to_highlight = sheet.cell(row=run_row, column=run_column)

    # Set the fill color to green
    cell_to_highlight.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Save the modified workbook
    workbook.save(r'\\cab-fs07.mae.virginia.edu\NewData\FieldTurf\2023-FieldTurf\1Video\Video Log Highlighted.xlsx')

def create_log():
    import shutil

    # Specify the paths for the original and copy Excel files
    original_file_path = r'\\cab-fs07.mae.virginia.edu\NewData\FieldTurf\2023-FieldTurf\1Video\Video Log.xlsx'
    copy_file_path = r'\\cab-fs07.mae.virginia.edu\NewData\FieldTurf\2023-FieldTurf\1Video\Video Log Highlighted.xlsx'

    # Use shutil to copy the file
    shutil.copyfile(original_file_path, copy_file_path)

